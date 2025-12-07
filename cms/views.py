from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from cms.models import Complaint, Category, Department, SLA, Escalation
from django.contrib.auth.models import User
from .models import StatusUpdate
from loginapp.models import UserProfile
from django.utils import timezone
from django.contrib.auth.models import User
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404
from .forms import ChangePasswordForm
from django.contrib.auth import update_session_auth_hash


@login_required
def user_dashboard(request):
    # Get user's complaints
    user_complaints = Complaint.objects.filter(user=request.user)

    # Dashboard stats
    total = user_complaints.count()
    open_count = user_complaints.filter(status="Pending").count()
    progress_count = user_complaints.filter(status="In Progress").count()
    closed_count = user_complaints.filter(status="Closed").count()

    # Recent complaints (last 5)
    recent = user_complaints.order_by('-created_at')[:5]

    return render(request, 'cms/user_dashboard.html', {
        "total": total,
        "open_count": open_count,
        "progress_count": progress_count,
        "closed_count": closed_count,
        "recent": recent,
    })


def admin_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)


@admin_required
def create_employee_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        email = request.POST.get("email")
        department_name = request.POST.get("department")
        level = request.POST.get("level")

        # Check duplicate username
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken!")
            return redirect("cms:create_employee")

        # Validate required fields
        if not department_name:
            messages.error(request, "Department is required!")
            return redirect("cms:create_employee")

        if not level:
            messages.error(request, "Level is required!")
            return redirect("cms:create_employee")

        # Create user
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email
        )

        # Ensure profile exists
        if not hasattr(user, 'profile'):
            UserProfile.objects.create(user=user)

        # Update the profile role to employee, department, level
        user.profile.role = "employee"
        user.profile.department = department_name
        user.profile.level = level
        user.profile.save()

        messages.success(request, "Employee created successfully!")

        return redirect('cms:create_employee')

    return render(request, 'admin/create_employee.html')


@login_required
def employee_dashboard(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != "employee":
        return redirect('login')  # unauthorized

    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')

    complaints = Complaint.objects.filter(assigned_to=request.user).select_related('user')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    complaints = complaints.order_by('-created_at')

    return render(request, 'employee/dashboard.html', {
        "complaints": complaints,
        "status_filter": status_filter,
        "assigned_filter": assigned_filter,
    })


@admin_required
def admin_dashboard(request):
    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')

    # Get all complaints for stats (unfiltered)
    all_complaints = Complaint.objects.all()

    # Dashboard stats (unfiltered)
    total = all_complaints.count()
    open_count = all_complaints.filter(status="Pending").count()
    progress_count = all_complaints.filter(status="In Progress").count()
    closed_count = all_complaints.filter(status="Closed").count()
    pending_approvals_count = all_complaints.filter(status="Resolved - Pending Approval").count()

    # Filtered complaints for recent display
    complaints = Complaint.objects.all().select_related('user', 'assigned_to')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    # Recent complaints (last 5, filtered)
    recent = complaints.order_by('-created_at')[:5]

    # Get escalated grievances
    escalated_grievances = Complaint.objects.filter(status="Escalated").select_related('user', 'assigned_to').order_by('-created_at')

    return render(request, 'admin/dashboard.html', {
        "total": total,
        "open_count": open_count,
        "progress_count": progress_count,
        "closed_count": closed_count,
        "pending_approvals_count": pending_approvals_count,
        "recent": recent,
        "escalated_grievances": escalated_grievances,
        "status_filter": status_filter,
        "assigned_filter": assigned_filter,
    })

@admin_required
def assign_grievance(request):
    if request.method == "POST":
        complaint_id = request.POST.get("complaint")
        department = request.POST.get("department")
        category = request.POST.get("category")
        employee_id = request.POST.get("employee")

        complaint = Complaint.objects.get(complaint_id=complaint_id)
        employee = User.objects.get(id=employee_id)

        # Validate that the employee is in the selected department
        if hasattr(employee.profile, 'department') and employee.profile.department and employee.profile.department != department:
            messages.error(request, "Selected employee is not in the selected department!")
            return redirect('cms:assign_grievance')

        complaint.assigned_to = employee
        complaint.status = "Assigned"
        complaint.save()

        remarks = f"Assigned to {employee.username} in {department}"
        if category:
            remarks += f" - Category: {category}"

        StatusUpdate.objects.create(
            complaint=complaint,
            status="Assigned",
            remarks=remarks
        )

        messages.success(request, "Grievance assigned successfully!")
        return redirect('cms:assign_grievance')

    # Get unassigned complaints and employees
    unassigned_complaints = Complaint.objects.filter(assigned_to__isnull=True)
    employees = User.objects.filter(profile__role="employee")

    # Prepare complaints data for JSON serialization
    complaints_data = {}
    for complaint in unassigned_complaints:
        complaints_data[complaint.complaint_id] = {
            'subject': complaint.subject,
            'details': complaint.details,
            'status': complaint.status,
            'user': complaint.user.username,
            'created_at': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        }

    return render(request, 'admin/assign_grievance.html', {
        'complaints': unassigned_complaints,
        'employees': employees,
        'complaints_data': complaints_data,
    })
    

@login_required
def update_complaint_status(request, complaint_id):
    complaint = Complaint.objects.get(pk=complaint_id)

    if request.user != complaint.assigned_to:
        return redirect('employee_dashboard')

    if request.method == "POST":
        status = request.POST.get("status")
        remarks = request.POST.get("remarks")

        # Check if complaint is being resolved and has high/critical priority
        if status == "Resolved" and complaint.category and complaint.category.priority in ['High', 'Critical']:
            status = "Resolved - Pending Approval"

        complaint.status = status
        complaint.save()

        StatusUpdate.objects.create(
            complaint=complaint,
            status=status,
            remarks=remarks
        )

        return redirect('cms:employee_dashboard')

    return render(request, "employee/update_status.html", {"complaint": complaint})

@login_required
def new_complaint(request):
    if request.method == "POST":
        department = request.POST.get("department")
        subject = request.POST.get("subject")
        details = request.POST.get("details")

        Complaint.objects.create(
            user=request.user,
            category=None,
            department=department,
            subject=subject,
            details=details,
            status="Pending"
        )

        messages.success(request, "Complaint submitted successfully!")
        return redirect('cms:new_complaint')

    return render(request, "user/new_complaint.html")


@login_required
def my_complaints(request):
    search = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    complaints = Complaint.objects.filter(user=request.user)

    if search:
        complaints = complaints.filter(subject__icontains=search)

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    return render(request, "user/my_complaints.html", {
        "complaints": complaints,
        "search": search,
        "status_filter": status_filter
    })

@login_required
def complaint_timeline(request, id):
    complaint = Complaint.objects.get(complaint_id=id, user=request.user)
    updates = StatusUpdate.objects.filter(complaint=complaint).order_by('date')

    return render(request, "user/timeline.html", {
        "complaint": complaint,
        "updates": updates
    })


@login_required
def update_complaint(request, complaint_id):
    complaint = Complaint.objects.get(complaint_id=complaint_id, user=request.user)

    if complaint.status != "Pending":
        messages.error(request, "You can only update pending complaints.")
        return redirect('cms:my_complaints')

    categories = Category.objects.all()

    if request.method == "POST":
        category_id = request.POST.get("category")
        subject = request.POST.get("subject")
        details = request.POST.get("details")

        category = Category.objects.get(category_id=category_id)

        complaint.category = category
        complaint.subject = subject
        complaint.details = details
        complaint.save()

        messages.success(request, "Complaint updated successfully!")
        return redirect('cms:my_complaints')

    return render(request, "user/update_complaint.html", {
        "complaint": complaint,
        "categories": categories
    })

@login_required
def withdraw_complaint(request, complaint_id):
    complaint = Complaint.objects.get(complaint_id=complaint_id, user=request.user)

    if complaint.status != "Pending":
        messages.error(request, "You can only withdraw pending complaints.")
        return redirect('cms:my_complaints')

    complaint.status = "Withdrawn"
    complaint.save()

    StatusUpdate.objects.create(
        complaint=complaint,
        status="Withdrawn",
        remarks="Complaint withdrawn by user"
    )

    messages.success(request, "Complaint withdrawn successfully!")
    return redirect('cms:my_complaints')

@login_required
def approve_resolution(request, complaint_id):
    complaint = Complaint.objects.get(complaint_id=complaint_id, user=request.user)

    if complaint.status != "Resolved":
        messages.error(request, "Only resolved complaints can be approved.")
        return redirect('cms:my_complaints')

    complaint.status = "Closed"
    complaint.save()

    StatusUpdate.objects.create(
        complaint=complaint,
        status="Closed",
        remarks="Resolution approved by user"
    )

    messages.success(request, "Resolution approved successfully!")
    return redirect('cms:my_complaints')

@login_required
def reopen_complaint(request, complaint_id):
    complaint = Complaint.objects.get(complaint_id=complaint_id, user=request.user)

    if complaint.status != "Resolved":
        messages.error(request, "Only resolved complaints can be reopened.")
        return redirect('cms:my_complaints')

    complaint.status = "Reopened"
    complaint.save()

    StatusUpdate.objects.create(
        complaint=complaint,
        status="Reopened",
        remarks="Complaint reopened by user"
    )

    messages.success(request, "Complaint reopened successfully!")
    return redirect('cms:my_complaints')


@login_required
def edit_profile(request):
    profile = request.user.profile

    if request.method == "POST":
        request.user.first_name = request.POST.get("first_name")
        request.user.last_name = request.POST.get("last_name")
        profile.phone = request.POST.get("phone")

        request.user.save()
        profile.save()

        return redirect("cms:user_dashboard")

    return render(request, "user/edit_profile.html", {
        "profile": profile
    })


@login_required
def user_report_pdf(request):
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.filter(user=request.user)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="complaints.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(f"Your Complaint Report - {request.user.username}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Prepare table data
    data = [['ID', 'Subject', 'Department', 'Status', 'Created']]

    for c in complaints:
        created_date = c.created_at.strftime('%Y-%m-%d %H:%M')
        data.append([
            str(c.complaint_id),
            c.subject,
            c.department,
            c.status,
            created_date
        ])

    # Create table
    table = Table(data)

    # Add table style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def user_report_excel(request):
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.filter(user=request.user)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

<<<<<<< HEAD
    # Add title
    ws['A1'] = f"Your Complaint Report - {request.user.username}"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)

    # Add headers
    headers = ["ID", "Subject", "Department", "Status", "Created"]
    ws.append([])
    ws.append(headers)

    # Make headers bold
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = cell.font.copy(bold=True)

    # Add data
    for c in complaints:
        created_date = c.created_at.strftime('%Y-%m-%d %H:%M')
        ws.append([
            c.complaint_id,
            c.subject,
            c.department,
            c.status,
            created_date
        ])

    # Set column widths
    column_widths = [10, 30, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
=======
    ws.append(["ID", "Subject", "Status", "Created"])

    for c in complaints:
        ws.append([c.complaint_id, c.subject, c.status, str(c.created_at)])
>>>>>>> e2f363d1db38133cafa260c091eb4d546218ad97

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="complaints.xlsx"'
    wb.save(response)
    return response

@admin_required
def admin_report_pdf(request):
    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.all().select_related('user', 'assigned_to')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="admin_complaints.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("Admin Complaint Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Prepare table data
    data = [['ID', 'User', 'Subject', 'Department', 'Status', 'Assigned To', 'Created']]

    for c in complaints:
        assigned = c.assigned_to.username if c.assigned_to else "Unassigned"
        created_date = c.created_at.strftime('%Y-%m-%d %H:%M')
        data.append([
            str(c.complaint_id),
            c.user.username,
            c.subject,
            c.department,
            c.status,
            assigned,
            created_date
        ])

    # Create table
    table = Table(data)

    # Add table style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    table.setStyle(style)

    elements.append(table)
    doc.build(elements)
    return response


@admin_required
def admin_report_excel(request):
    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.all().select_related('user', 'assigned_to')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

    ws.append(["ID", "User", "Subject", "Department", "Status", "Assigned To", "Created"])

    for c in complaints:
        assigned = c.assigned_to.username if c.assigned_to else "Unassigned"
        ws.append([c.complaint_id, c.user.username, c.subject, c.department, c.status, assigned, str(c.created_at)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="admin_complaints.xlsx"'
    wb.save(response)
    return response

@login_required
def employee_report_pdf(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != "employee":
        return redirect('login')

    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.filter(assigned_to=request.user).select_related('user')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="employee_complaints.pdf"'

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, f"Employee Complaint Report - {request.user.username}")
    y -= 40

    p.setFont("Helvetica", 12)

    for c in complaints:
        p.drawString(50, y, f"#{c.complaint_id} | {c.subject} | {c.status} | {c.user.username}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.showPage()
    p.save()
    return response


@login_required
def employee_report_excel(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != "employee":
        return redirect('login')

    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.filter(assigned_to=request.user).select_related('user')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

    ws.append(["ID", "User", "Subject", "Department", "Status", "Created"])

    for c in complaints:
        ws.append([c.complaint_id, c.user.username, c.subject, c.department, c.status, str(c.created_at)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="employee_complaints.xlsx"'
    wb.save(response)
    return response

@login_required
def escalate_complaint(request, complaint_id):
    if not hasattr(request.user, 'profile') or request.user.profile.role != "employee":
        return redirect('login')

    complaint = get_object_or_404(Complaint, pk=complaint_id)

    # Check if the complaint is assigned to the current employee
    if complaint.assigned_to != request.user:
        messages.error(request, "You can only escalate complaints assigned to you.")
        return redirect('cms:employee_dashboard')

    # Create escalation record
    Escalation.objects.create(
        complaint=complaint,
        escalated_to_level="Level-2",
        reason="Escalated by employee to Level 2"
    )

    # Update complaint status to Escalated and unassign from employee
    complaint.status = "Escalated"
    complaint.assigned_to = None
    complaint.save()

    # Create status update
    StatusUpdate.objects.create(
        complaint=complaint,
        status="Escalated",
        remarks="Escalated to Level 2 by employee"
    )

    messages.success(request, f"Complaint #{complaint.complaint_id} escalated to Level 2.")
    return redirect('cms:employee_dashboard')

@login_required
def employee_assignment_history(request):
    if not hasattr(request.user, 'profile') or request.user.profile.role != "employee":
        return redirect('login')

    # Get all status updates for complaints assigned to this employee
    assignment_history = StatusUpdate.objects.filter(
        complaint__assigned_to=request.user
    ).select_related('complaint', 'complaint__user').order_by('-date')

    return render(request, 'employee/assignment_history.html', {
        'assignment_history': assignment_history,
    })

@admin_required
def employee_management(request):
    employees = User.objects.filter(profile__role="employee")
    return render(request, 'admin/employee_management.html', {'employees': employees})

@admin_required
def view_employees(request):
    employees = User.objects.filter(profile__role="employee")
    return render(request, 'admin/view_employees.html', {'employees': employees})

@admin_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(User, id=employee_id, profile__role="employee")

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        department_name = request.POST.get("department")
        level = request.POST.get("level")

        # Check if username is taken by another user
        if User.objects.filter(username=username).exclude(id=employee.id).exists():
            messages.error(request, "Username already taken!")
            return redirect("cms:edit_employee", employee_id=employee.id)

        # Validate required fields
        if not department_name:
            messages.error(request, "Department is required!")
            return redirect("cms:edit_employee", employee_id=employee.id)

        if not level:
            messages.error(request, "Level is required!")
            return redirect("cms:edit_employee", employee_id=employee.id)

        employee.username = username
        employee.email = email
        employee.save()

        # Update department and level
        employee.profile.department = department_name
        employee.profile.level = level
        employee.profile.save()

        messages.success(request, "Employee updated successfully!")
        return redirect('cms:employee_management')

    return render(request, 'admin/edit_employee.html', {'employee': employee})

@admin_required
def delete_employee(request, employee_id):
    employee = get_object_or_404(User, id=employee_id, profile__role="employee")

    if request.method == "POST":
        employee.delete()
        messages.success(request, "Employee deleted successfully!")
        return redirect('cms:employee_management')

    return render(request, 'admin/delete_employee.html', {'employee': employee})

@admin_required
def view_users(request):
    role_filter = request.GET.get('role', '')
    department_filter = request.GET.get('department', '')

    users = User.objects.all().select_related('profile')

    if role_filter:
        if role_filter == 'admin':
            users = users.filter(is_superuser=True)
        else:
            users = users.filter(profile__role=role_filter)

    if department_filter:
        users = users.filter(profile__department=department_filter)

    # Ensure all users have profiles
    for user in users:
        if not hasattr(user, 'profile'):
            UserProfile.objects.create(user=user)

    return render(request, 'admin/view_users.html', {
        'users': users,
        'role_filter': role_filter,
        'department_filter': department_filter
    })

@admin_required
def view_all_complaints(request):
    status_filter = request.GET.get('status', '')
    assigned_filter = request.GET.get('assigned', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')

    complaints = Complaint.objects.all().select_related('user', 'assigned_to')

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if assigned_filter:
        if assigned_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        elif assigned_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__gte=start_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__lte=end_date)
        except ValueError:
            pass  # Invalid date, ignore filter

    complaints = complaints.order_by('-created_at')

    return render(request, 'admin/view_all_complaints.html', {
        'complaints': complaints,
        'status_filter': status_filter,
        'assigned_filter': assigned_filter,
        'start_date_filter': start_date_str,
        'end_date_filter': end_date_str
    })


@login_required
def change_password(request):
    if request.method == "POST":
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, "Password changed successfully!")
            return redirect('cms:user_dashboard')
    else:
        form = ChangePasswordForm(request.user)
    return render(request, 'user/change_password.html', {'form': form})

@admin_required
def delete_complaint(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)

    if request.method == "POST":
        # Create a status update for the user
        StatusUpdate.objects.create(
            complaint=complaint,
            status="Deleted",
            remarks="admin deleted it due to insufficient or incorrect data"
        )
        # Delete the complaint
        complaint.delete()
        messages.success(request, "Complaint deleted successfully!")
        return redirect('cms:view_all_complaints')

    return render(request, 'admin/delete_complaint.html', {'complaint': complaint})

@admin_required
def escalate_grievance(request):
    if request.method == "POST":
        complaint_id = request.POST.get("complaint")
        department = request.POST.get("department")
        category = request.POST.get("category")
        employee_id = request.POST.get("employee")

        complaint = Complaint.objects.get(complaint_id=complaint_id)
        employee = User.objects.get(id=employee_id)

        # Validate that the employee is Level 2
        if not hasattr(employee.profile, 'level') or employee.profile.level != "Level-2":
            messages.error(request, "Selected employee is not a Level 2 employee!")
            return redirect('cms:escalate_grievance')

        # Validate that the employee is in the selected department
        if hasattr(employee.profile, 'department') and employee.profile.department and employee.profile.department != department:
            messages.error(request, "Selected employee is not in the selected department!")
            return redirect('cms:escalate_grievance')

        complaint.assigned_to = employee
        complaint.status = "Assigned"
        complaint.save()

        remarks = f"Escalated and assigned to Level 2 employee {employee.username} in {department}"
        if category:
            remarks += f" - Category: {category}"

        StatusUpdate.objects.create(
            complaint=complaint,
            status="Assigned",
            remarks=remarks
        )

        messages.success(request, "Escalated grievance assigned successfully!")
        return redirect('cms:escalate_grievance')

    # Get escalated complaints
    escalated_complaints = Complaint.objects.filter(status="Escalated")
    # Get Level 2 employees
    level2_employees = User.objects.filter(profile__role="employee", profile__level="Level-2")

    # Prepare complaints data for JSON serialization
    complaints_data = {}
    for complaint in escalated_complaints:
        complaints_data[complaint.complaint_id] = {
            'subject': complaint.subject,
            'details': complaint.details,
            'status': complaint.status,
            'user': complaint.user.username,
            'created_at': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        }

    return render(request, 'admin/escalate_grievance.html', {
        'complaints': escalated_complaints,
        'employees': level2_employees,
        'complaints_data': complaints_data,
    })


@admin_required
def manage_grievances(request):
    if request.method == "POST":
        complaint_id = request.POST.get("complaint_id")
        action = request.POST.get("action")

        if not complaint_id or not complaint_id.isdigit():
            messages.error(request, "Invalid complaint ID.")
            return redirect('cms:manage_grievances')

        complaint = get_object_or_404(Complaint, pk=int(complaint_id))

        if action == "approve":
            complaint.status = "Closed"
            remarks = "Resolution approved by admin"
        elif action == "reject":
            complaint.status = "Reopened"
            remarks = "Resolution rejected by admin"
        elif action == "escalate":
            employee_id = request.POST.get("employee_id")
            if not employee_id or not employee_id.isdigit():
                messages.error(request, "Please select a level 2 employee to assign.")
                return redirect('cms:manage_grievances')
            employee = get_object_or_404(User, id=int(employee_id), profile__role="employee", profile__level="Level-2")

            # Assign to employee
            complaint.assigned_to = employee
            complaint.status = "Assigned"

            # Escalate to level 2
            Escalation.objects.create(
                complaint=complaint,
                escalated_to_level="Level-2",
                reason="SLA breached - escalated by admin"
            )
            remarks = f"Escalated to Level 2 and assigned to {employee.username} due to SLA breach"
        else:
            messages.error(request, "Invalid action.")
            return redirect('cms:manage_grievances')

        complaint.save()

        StatusUpdate.objects.create(
            complaint=complaint,
            status=complaint.status,
            remarks=remarks
        )

        messages.success(request, f"Grievance {complaint.complaint_id} {action}ed successfully!")
        return redirect('cms:manage_grievances')

    # Get resolved complaints (including pending approval)
    resolved_complaints = Complaint.objects.filter(status__in=["Resolved", "Resolved - Pending Approval"])

    # Get SLA breaches
    sla_breaches = SLA.objects.filter(breached=True).select_related('complaint')

    # Get level 2 employees
    level2_employees = User.objects.filter(profile__role="employee", profile__level="Level-2")

    return render(request, 'admin/manage_grievances.html', {
        'complaints': resolved_complaints,
        'sla_breaches': sla_breaches,
        'level2_employees': level2_employees
    })


